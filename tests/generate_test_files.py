"""
Reads original_data files and removes real data so they can be uploaded to git
"""
import os
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell import Cell, MergedCell

from liquidaciones.conciliation_model import Conciliation
from liquidaciones.openpyxl_helpers import df_to_excel

_DEBUG = False


# _DEBUG = True


def exact_match(pattern: str) -> str:
    """Returns regular expression that match exactly the given string. Adds ^ for match at beginning and $ for match to
    the end"""
    return f"^{pattern}$"


def to_list(value) -> list | None:
    """Converts a value to a list if is not None or already a list"""
    if value is None or isinstance(value, list):
        return value
    else:
        return [value]


def str_cell_value(cell: Cell | MergedCell) -> str:
    """Gets cell value as string. If empty returns empty string, otherwise cell value converted to string"""
    if not cell.value:
        return ""
    else:
        return str(cell.value)


def col2int(col_name: str) -> int:
    """Converts a col name to a string"""
    if len(col_name) > 1:
        raise NotImplemented("Conversion of columns of more than 1 char is not implemented")
    return ord(col_name.upper()) - ord("A")


class CellsFilter:
    # Markers of last row
    TILL_END = -1
    TILL_BLANK = None

    def __init__(self, cells: list, first_col: int | list, first_row: int | list, last_row: int | str = None,
                 include_header: bool = False):
        """
        Creates a configuration for a cell selector
        Examples:
            CellsSelector(cells, first_col=0, first_row=["sample header"], last_row=-1, include_header=False)
            CellsSelector(cells, first_col=0, first_row=["sample header"], last_row=-1)
                Looks for data in column 0, starting in the row bellow where a column that has sample header inside
                is found
        Args:
            cells: a list of lists of cells, e.g. cells[row][col]. Is the output of list(WorkBook.cells) of openpyxl
            first_col: a way to calculate the first column where data to be selected starts. It can be either a number
             (0-based) or a list (a list of regular expressions to match with the cell in the first row)
            first_row: a way to calculate the first row where data to be selected starts. It can be either an integer
             (0-based) for selecting data or a list of regular expressions to find each column of data.
            last_row: Can be a number of row (0-based) for end of data selection (use -1 or CellsSelector.TILL_END to
            read until the end of the sheet), None (CellsSelector.TILL_BLANK) to read until a blank cell or a string
            with a regular expression to match for the last row. Defaults to None
            include_header: False to skip header (default value), True to include header in selection
        """
        self.cells = cells
        self.first_col = first_col
        self.first_row = first_row
        self.last_row = last_row
        self.include_header = include_header

        # Check data
        if self.first_row is None and self.first_col is None:
            raise ValueError("Either first_row or first_col must be informed")
        if not (isinstance(self.first_row, int) or isinstance(self.first_col, int)):
            raise ValueError("Either first_row or first_col must be an integer")

    def filter(self) -> list:
        if isinstance(self.first_col, list):
            filter_cols = self.get_cols_matching_first_col(self.cells[self.first_row])
            start_row = self.first_row
        else:
            search_pattern = to_list(self.first_row)
            start_row, filter_cols = self.get_rows_cols_matching_pattern(search_pattern,
                                                                         start_col=self.first_col)
            if start_row is None:
                # Not found. Use brute force searching in all columns, it might start earlier
                start_row, filter_cols = self.get_rows_cols_matching_pattern(search_pattern)
                if start_row is None:
                    # Still not found even using brute force. Pattern is not present in the row. Return empty value
                    return list(list() for _ in range(len(search_pattern)))

        # Start row might include header. Skip it according to configuration
        if not self.include_header:
            start_row += 1

        end_row = self.get_last_row(filter_cols[0], start_row)

        filter_cells = []
        for i_col, col in enumerate(filter_cols):
            filter_cells.append([])
            for row in range(start_row, end_row):
                filter_cells[i_col].append(self.cells[row][col])

        return filter_cells

    def get_rows_cols_matching_pattern(self, pattern_list: list, start_row=0, end_row=-1,
                                       start_col=0) -> tuple:
        """The mega finder: looks for any row that matches all values in pattern list"""
        if end_row == -1:
            end_row = len(self.cells) - 1
        for idx_row in range(start_row, end_row + 1):
            row = self.cells[idx_row]
            cols_found = []
            pattern_found = []
            for idx_col in range(start_col, len(row)):
                cell = row[idx_col]
                for pattern in pattern_list:
                    if pattern in pattern_found:
                        continue
                    if re.findall(pattern, str_cell_value(cell)):
                        # If first pattern is not found before any other pattern, skip row
                        if pattern != pattern_list[0] and pattern_list[0] not in pattern_found:
                            continue
                        cols_found.append(idx_col)
                        pattern_found.append(pattern)
                        break
                if len(cols_found) == len(pattern_list):
                    # All patterns found in current row
                    return idx_row, list(cols_found)
        return None, list()

    def get_cols_matching_first_col(self, cells_first_row: list) -> list:
        """Gets a list of col indexes in the first row, matching in self.first_col"""
        cols = list()
        for idx_col, cell in enumerate(cells_first_row):
            for pattern in self.first_col:
                if _DEBUG:
                    print(idx_col, cell.value)
                if re.findall(pattern, str_cell_value(cell)):
                    cols.append(idx_col)
                    if len(cols) == len(self.first_col):
                        return cols
        return cols

    def get_last_row(self, search_column: int, start_row: int) -> int:
        """Calculates last row as an integer"""
        if self.last_row == self.TILL_END:
            return len(self.cells)
        elif isinstance(self.last_row, int):
            return self.last_row + 1  # Otherwise range does not work properly
        else:
            # The value must be found in the search_column
            if self.last_row is self.TILL_BLANK:
                pattern = exact_match("")
            elif isinstance(self.last_row, str):
                pattern = self.last_row
            else:
                raise ValueError(f"Value for last row cannot be processed: last_row={self.last_row}")
            # Find pattern in the search column
            for idx_row in range(start_row, len(self.cells)):
                if re.findall(pattern, str_cell_value(self.cells[idx_row][search_column])):
                    return idx_row
            return len(self.cells)  # If not found return all available data (index of last row of data)
        pass

    def get_row_matching_first_row(self, cells: list, pattern_search: str) -> int:
        """Gets the row that matches the regular expression of self.first_row in the column self.first_col"""
        for idx_row, row in enumerate(cells):
            if re.findall(pattern_search, str_cell_value(row[self.first_col])):
                return idx_row


class FilePrivateDataCleaner:
    """
    Class that reads a file and removes all private data
    """

    def __init__(self, filename, suffix: str = None):
        self.__wb = load_workbook(filename)
        self.__old_filename = filename
        new_filename = os.path.basename(filename).replace("_original", "")
        if suffix:
            new_filename = suffix.join(os.path.splitext(new_filename))
        save_dir = os.path.join(os.path.dirname(__name__), "../data/test_data")
        os.makedirs(save_dir, exist_ok=True)
        self.__new_filename = os.path.join(save_dir, new_filename)
        self.__sheet_names = self.__wb.sheetnames
        # Reads all cells from all sheets (slow but useful)
        self.__cells = {s: list(self.__wb[s].rows) for s in self.__sheet_names}
        print("Old: ", self.__old_filename)
        print("New: ", self.__new_filename)

    def __save(self):
        # Remove properties from document. First, check if there is any
        for prop in self.__wb.custom_doc_props.props:
            print(f"{prop.name}: {prop.value}")
        self.__wb.save(self.__new_filename)

    def __getitem__(self, sheet_name):
        if sheet_name in self.__sheet_names:
            return self.__cells[sheet_name]
        else:
            raise ValueError(f"Sheet {sheet_name} not found in {self.__old_filename}")

    def process(self, pattern: str | list, sheet_config: list):
        cells = None
        for cfg in sheet_config:
            new_cells = cfg.filter()
            if not cells:
                cells = new_cells
            else:
                for col in range(len(new_cells)):
                    cells[col].extend(new_cells[col])
            if _DEBUG:
                print(new_cells)
                if new_cells:
                    print(len(new_cells[0]))

        idx_fmt = "_{idx}"

        n_cols = len(cells)
        if pattern == "":
            new_pattern = ["" for _ in range(n_cols)]
        elif isinstance(pattern, str):
            new_pattern = [pattern + idx_fmt for _ in range(n_cols)]
        elif isinstance(pattern, list):
            if len(pattern) != n_cols:
                raise ValueError("The number of patterns does not match the number of columns found")
            new_pattern = [p + idx_fmt for p in pattern]
        elif pattern is None:
            raise NotImplementedError("Caso None no contemplado")
        else:
            raise NotImplementedError("Unknown case")

        pattern = to_list(new_pattern)
        if cells:
            for col in range(n_cols):
                single_values = list(dict.fromkeys(c.value for c in cells[col] if c.value))
                single_dict = {value: pattern[col].format(idx=idx + 1) for idx, value in enumerate(single_values)}
                for c in cells[col]:
                    c.value = single_dict.get(c.value)
                pass

    def remove_all_images(self, sheet_name: str):
        """Deletes all images in given sheet"""
        ws = self.__wb[sheet_name]
        ws._images = []

    def clear_all_cells(self, sheet_name: str):
        for row in self[sheet_name]:
            for cell in row:
                if not isinstance(cell, MergedCell):
                    cell.value = None

    def process_global_file(self):
        """Process bank data from a global file"""

        def make_config_header(sheet, header):
            """Shortcut function for searching a header in the first row of a sheet"""
            return CellsFilter(self[sheet], first_col=[exact_match(header)], first_row=0, last_row=CellsFilter.TILL_END)

        ###################################
        # process bank data (just concepto)
        ###################################
        self.process(pattern="Concepto_Banco_fake",
                     sheet_config=[
                         make_config_header(sheet="banco", header="Concepto"),
                         make_config_header(sheet="banco_ingresos", header="Concepto"),
                         make_config_header(sheet="banco_gastos", header="Concepto_0"),
                     ])
        ###################################
        # process income data
        ###################################
        self.process(pattern="Piso_fake",
                     sheet_config=[
                         make_config_header(sheet="ingresos", header="Piso/Local"),
                         make_config_header(sheet="banco_ingresos", header="Piso/Local"),
                     ])
        self.process(pattern="Inquilino_fake",
                     sheet_config=[
                         make_config_header(sheet="ingresos", header="Inquilino"),
                         make_config_header(sheet="banco_ingresos", header="Inquilino"),
                     ])
        ###################################
        # process expenses data
        ###################################
        self.process(pattern="Concepto_Gastos_fake",
                     sheet_config=[
                         make_config_header(sheet="gastos", header="CONCEPTO"),
                         make_config_header(sheet="banco_gastos", header="CONCEPTO"),
                     ])

        ###################################
        # process "finca" column
        ###################################
        self.process(pattern="Finca_fake",
                     sheet_config=[
                         make_config_header(sheet="ingresos", header="finca"),
                         make_config_header(sheet="gastos", header="finca"),
                         make_config_header(sheet="banco_ingresos", header="finca"),
                         make_config_header(sheet="banco_gastos", header="finca"),
                     ])
        self.__save()
        return

    def process_bank_extract(self):
        """Process bank file from a bank extract file"""

        # Delete any sheet but sheet 0
        for ws in self.__wb.worksheets[1:]:
            self.__wb.remove(ws)
        sheet_name = self.__sheet_names[0]
        ##################################
        # process just the first sheet
        ##################################
        for col, marker in [(2, "Cuenta"), (2, "Titular"), (3, "Saldo disponible"), (3, "Retenciones"),
                            (4, "Saldo real"), (4, "Saldo consolidado")]:
            self.process(pattern=marker + "_Fake",
                         sheet_config=[
                             CellsFilter(self[sheet_name], first_col=col, first_row=[marker])
                         ])
        self.process(pattern="Concepto_Fake",
                     sheet_config=[CellsFilter(self[sheet_name], first_col=['Concepto'], first_row=7)])
        self.process(pattern="",
                     sheet_config=[CellsFilter(self[sheet_name], first_col=['Saldo'], first_row=7)])
        self.remove_all_images(sheet_name)
        self.__save()

    def process_accounting(self):
        """Processes a Gesfincas output file"""
        sheets = self.__sheet_names[:-1]  # Skip last sheet (as it is a summary)
        self.process(sheet_config=[
            CellsFilter(cells=self[s], first_col=2, first_row=["Piso/Local", "Inquilino"], last_row="Total Finca: ")
            for s in sheets], pattern=["Piso_fake", "Inquilino_fake"])
        self.process(sheet_config=[
            CellsFilter(cells=self[s], first_col=col2int("B"), first_row=["CONCEPTO"], last_row="Total Finca: ")
            for s in sheets], pattern="CONCEPTO")
        self.process(sheet_config=[
            CellsFilter(cells=self[s], first_col=["Finca: "], first_row=7, last_row=7, include_header=True)
            for s in sheets], pattern="Finca: Finca_Fake")
        self.process(sheet_config=[
            CellsFilter(cells=self[s], first_col=["NIF: "], first_row=5, last_row=5, include_header=True)
            for s in sheets], pattern="NIF: NIF_Fake")
        self.process(sheet_config=[
            CellsFilter(cells=self[s], first_col=['MADRID'], first_row=3, last_row=3, include_header=True)
            for s in sheets], pattern="Propietario_Fake")
        self.process(sheet_config=[
            CellsFilter(cells=self[s], first_col=col2int("B"), first_row=["Total Finca: "], include_header=True)
            for s in sheets], pattern="Total Finca: Finca_Fake")

        for s in self.__sheet_names:
            self.remove_all_images(s)
        # Clear contents of last sheet
        self.clear_all_cells(self.__sheet_names[-1])
        self.__save()

    def generate_global_test_data(self):
        """Generates test data from original_data file"""
        in_dfs = {}
        with pd.ExcelFile(self.__old_filename) as xls:
            for s in xls.sheet_names:
                in_dfs[s] = pd.read_excel(xls, sheet_name=s)

        # Make changes needed for testing
        df_bank = in_dfs['banco']
        df_exp = in_dfs['gastos']
        """Changes are:        
        0,  # Deleted row from expenses
        1,  # Modified cash value in bank (multiplied by 1000)
        2,  # Deleted row from bank
        3,  # Modified cash value in expenses (multiplied by 10)
        73, 74,  # These two rows are unchanged, but as they have exactly the same values, cannot bucket them
        220,  # Deleted one row from bank (there are multiple rows)
        210,  # Deleted one row from expenses (there are two rows)
        """
        col_bucket = Conciliation.col_bucket
        factor = 1000
        df_exp.drop(index=df_exp[df_exp[col_bucket] == 0].index, inplace=True)
        df_bank.loc[df_bank[col_bucket] == 1, "Importe"] *= factor
        df_bank.drop(index=df_bank[df_bank[col_bucket] == 2].index, inplace=True)
        df_exp.loc[df_exp[col_bucket] == 3, 'Pagos'] *= factor
        df_bank.drop(index=df_bank[df_bank[col_bucket] == 220].index[-1], inplace=True)
        df_exp.drop(index=df_exp[df_exp[col_bucket] == 210].index[-1], inplace=True)

        with pd.ExcelWriter(self.__new_filename) as out_xls:
            for s, df in in_dfs.items():
                df_to_excel(df, out_xls, s)


if __name__ == '__main__':
    FilePrivateDataCleaner("../data/original_data/global_test_data_original.xlsx").process_global_file()
    FilePrivateDataCleaner("../data/original_data/liquidaciones_original.xlsx").process_accounting()
    FilePrivateDataCleaner("../data/original_data/MovimientosCuenta 16.7.2023_original.xlsx").process_bank_extract()
    FilePrivateDataCleaner("../data/test_data/global_test_data.xlsx", "_changed").generate_global_test_data()
